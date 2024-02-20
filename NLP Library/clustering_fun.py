import openpyxl
import re
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
import pickle
from bs4 import BeautifulSoup
import requests
import pandas as pd
import numpy as np
import math
from sklearn.metrics import silhouette_score
import skfuzzy as fuzz
import codecs
import gzip


# lowercase/case folding
def lowercase(data):
    new_data = []
    for i in range(len(data)):
        tempt = str(data[i]).lower()
        new_data.append(tempt.replace('b"suara.com - ', ""))
    return new_data


# remove symbol
def symbol_remover(data):
    new_data = []
    for i in range(len(data)):
        new_data.append(re.sub(r'[^\w]', ' ', data[i].replace("b suara com", "")))
    return new_data


# tokenisasi
def tokenisasi(data):
    new_data = []
    for i in range(len(data)):
        new_data.append(word_tokenize(data[i]))
    return new_data


# filter/stopwords
def stopword_s(data):
    # factory = StopWordRemoverFactory()
    # stopword = factory.create_stop_word_remover()
    new_data = []
    stop_words = set(stopwords.words('indonesian'))
    for i in range(len(data)):
        tempt = []
        for word in data[i]:
            if word not in stop_words:
                tempt.append(word)
        new_data.append(tempt)
        # new_data.append(stopword.remove(data[i]))
    return new_data


# stemming
def stemming(data):
    factory = StemmerFactory()
    stemmer = factory.create_stemmer()

    new_data = []
    for i in range(len(data)):
        tempt = ""
        for word in data[i]:
            tempt = tempt + " " + word
        new_word = stemmer.stem(tempt)
        new_data.append(new_word.split())
    return new_data


# collecting term in all document
def collecting_fiture(data):
    new_data = []
    for i in range(len(data)):
        for word in data[i]:
            if word not in new_data:
                new_data.append(word)
    return new_data


# VSM
def vsm(data, term, wb):
    wb.create_sheet('vsm')
    sheet1 = wb['vsm']
    new_data = []
    # create term in excel
    i = 2
    for word in term:
        sheet1.cell(row=1, column=i).value = word
        i += 1

    # count term in all document
    dummy = 2
    for i in range(len(data)):
        sheet1.cell(row=i + dummy, column=1).value = i + 1
        tempt = []
        j = 2  # column in excel
        for word in term:
            tempt.append(data[i].count(word))
            sheet1.cell(row=i + dummy, column=j).value = data[i].count(word)
            j += 1
        new_data.append(tempt)
    return new_data


# save data
def create_excel():
    # Workbook is created
    # return xlwt.Workbook()
    return openpyxl.Workbook()


# add sheet
def add_sheet(sheet, data, wb):
    # add_sheet is used to create sheet.
    wb.create_sheet(sheet)
    sheet1 = wb[sheet]

    for i in range(len(data)):
        sheet1.cell(row=i + 1, column=1).value = i + 1
        sheet1.cell(row=i + 1, column=2).value = data[i]


# add sheet with data is list
def add_sheet_list(sheet, data, wb):
    # add_sheet is used to create sheet.
    wb.create_sheet(sheet)
    sheet1 = wb[sheet]

    for i in range(len(data)):
        sheet1.cell(row=i + 1, column=1).value = i + 1
        for j in range(len(data[i])):
            sheet1.cell(row=i + 1, column=j + 2).value = data[i][j]
        print(i)


# save
def save_excel(name, wb):
    wb.save('excel_logs/' + name + '.xlsx')


def save_list(name, l):
    with open("data_logs/" + name + ".txt", "wb") as fp:  # Pickling
        pickle.dump(l, fp)
    return print("Success to save....")


def load_list(name):
    with open("data_logs/" + name + ".txt", "rb") as fp:  # Unpickling
        # with codecs.open("data_logs/"+name+".txt", "rb", encoding="utf-8-sig") as fp:
        print("Hello", fp)
        b = pickle.load(fp)
    print("Success to load.....")
    return b


def crawl_web(url):
    new_url = url
    i = '?page='
    page = 1
    loop = True
    data = []
    while loop == True:
        # find article link
        req = requests.get(new_url)
        soup = BeautifulSoup(req.text, "lxml")
        news_links = soup.find_all("li", {'class': 'item-outer'})

        # check content available
        if not news_links:
            loop = False

        # looping through article link
        for idx, news in enumerate(news_links):
            try:
                # find news title
                title_news = news.find('a', {'class': 'ellipsis2'}).text

                # find urll news
                url_news = news.find('a', {'class': 'ellipsis2'}).get('href')

                # find news content in url
                req_news = requests.get(url_news)
                soup_news = BeautifulSoup(req_news.text, "lxml")

                # find news content
                news_content = soup_news.find("div", {'class': 'content-article'})

                # find paragraph in news content
                p = news_content.find_all('p')
                content = ' '.join(item.text for item in p)
                news_content = content.encode('utf8', 'replace')

                tempt = []
                tempt.append(title_news)
                tempt.append(news_content)
                tempt.append(url_news)

                data.append(tempt)
                # set_data(title_news,news_content,url_news)
                # print('tidak ada')

            except:
                pass
        print(page)
        # next pagination
        if (page > 50):
            loop = False
        page += 1
        new_url = url + i + str(page)

    return data


def save_content(data):
    woorkbook = create_excel()
    woorkbook.create_sheet('crawl_data')
    sheet1 = woorkbook['crawl_data']
    i = 0

    content = []
    for i in range(len(data)):
        for j in range(len(data[0])):
            sheet1.cell(row=i + 1, column=j + 1).value = data[i][j]
            if (j == 1):
                content.append(data[i][j])

    save_excel('crawl_data', woorkbook)
    save_list("data", content)


# n-gram


def generate_ngrams(data, n):
    # Use the zip function to help us generate n-grams
    # Concatentate the tokens into ngrams and return
    n_gram = []
    for i in range(0, len(data)):
        tokens = data[i]
        ngrams = zip(*[tokens[i:] for i in range(n)])
        n_gram.append([" ".join(ngram) for ngram in ngrams])
    return n_gram


def drop_highly_correlation(X, Y):
    # Convert feature matrix into DataFrame
    df = pd.DataFrame(X)

    # Create correlation matrix
    corr_matrix = df.corr().abs()
    # Select upper triangle of correlation matrix
    upper = corr_matrix.where(np.triu(np.ones(corr_matrix.shape), k=1).astype(np.bool))

    # Find index of feature columns with correlation greater than 0.95
    to_drop = [column for column in upper.columns if any(upper[column] > 0.95)]

    # Drop features
    new_data = df.drop(df.columns[to_drop], axis=1)

    header = list(new_data.columns.values)

    new_collecting_fiture = []
    for i in header:
        new_collecting_fiture.append(Y[i])

    return new_data, new_collecting_fiture


def tf(data, term, wb, sheet):
    wb.create_sheet(sheet)
    sheet1 = wb[sheet]
    new_data = []
    # create term in excel
    i = 2
    for word in term:
        sheet1.cell(row=1, column=i).value = word
        i += 1

    # count term in all document
    dummy = 2
    for i in range(len(data)):
        sheet1.cell(row=i + dummy, column=1).value = i + 1
        tempt = []
        j = 2  # column in excel
        for word in term:
            tempt.append(data[i].count(word))
            sheet1.cell(row=i + dummy, column=j).value = data[i].count(word) / len(data[i])
            j += 1
        new_data.append(tempt)
    return new_data


def idf(data):
    new_data = []

    for i in range(len(data[0])):
        count = 0
        for j in range(len(data)):
            count += int(data[j][i])
        new_data.append(math.log10(len(data) / count))
    return new_data


def tf_idf(tf, idf, term, wb):
    wb.create_sheet('tf_idf')
    sheet1 = wb['tf_idf']

    new_data = []

    i = 2
    for word in term:
        sheet1.cell(row=1, column=i).value = word
        i += 1

    # count term in all document
    dummy = 2

    for i in range(len(tf)):
        sheet1.cell(row=i + dummy, column=1).value = i + 1
        tempt = []
        for j in range(len(tf[i])):
            tempt.append(int(tf[i][j]) * int(idf[j]))
            sheet1.cell(row=i + dummy, column=j + dummy).value = float(tf[i][j]) * float(idf[j])
        new_data.append(tempt)
    return new_data


def clustering(data):
#    print(data)
    s_avg = []
    for i in range(2, len(data)):
#        print(i)
        cntr, u, u0, distant, fObj, iterasi, fpc = fuzz.cmeans(np.asarray(data).T, i, 2, 0.00001, 1000, seed=0)
        membership = np.argmax(u, axis=0)
#        print(u)

        # silhouette = silhouette_samples(tfidf, membership)
        s_avg.append(silhouette_score(data, membership, random_state=10))
    bestNumclusters = s_avg.index(max(s_avg)) + 2
    cntr, u, u0, distant, fObj, iterasi, fpc = fuzz.cmeans(np.asarray(data).T, bestNumclusters, 2, 0.00001, 1000, seed=0)
       
#    print(bestNumclusters)
    return u
